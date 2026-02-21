import openpyxl as xl
from openpyxl.chart import BarChart,Reference

Book = xl.load_workbook('Book1.xlsx')
Sheet1 = Book['Sheet1']

for row in range(2,Sheet1.max_row+1):
     cell = Sheet1.cell(row=row, column=3)
     corrected_price = cell.value*0.9
     corrected_price_cell = Sheet1.cell(row=row, column=4)
     corrected_price_cell.value = corrected_price
values = Reference(Sheet1,min_row = 2, max_row = Sheet1.max_row,min_col=4,max_col=4)

chart = BarChart()
chart.add_data(values)
Sheet1.add_chart(chart,'e2')

Book.save('Book3.xlsx')